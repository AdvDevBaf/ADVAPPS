from django.shortcuts import render
from .models import BlogPost,Category
from django.shortcuts import get_object_or_404, render_to_response, render
import requests
import pandas as pd
import os
import mimetypes
from django.http import HttpResponse
import datetime
import time
from django.views.decorators.csrf import csrf_exempt
from .getaddress import example
from django.http import JsonResponse, HttpResponse, HttpResponseRedirect
from tkinter.filedialog import askopenfilename
import openpyxl
from .forms import ArticleFileForm
import re

# Create your views here.


def index(request):
    entries = BlogPost.objects.published()
    return render_to_response('BafosApp/index.html', {'entries': entries})


def category(request, id, slug):
    category = get_object_or_404(Category, id=id, slug=slug)
    entries = category.entries.published()
    return render_to_response('BafosApp/index.html', {'entries': entries})


def archive_year(request, year):
    entries = BlogPost.objects.filter(draft=False,
                                      published_date__year=year)
    return render_to_response('BafosApp/index.html', {'entries': entries})


def archive_month(request, year, month):
    entries = BlogPost.objects.filter(draft=False,
                                      published_date__year=year,
                                      published_date__month=month)
    return render_to_response('BafosApp/index.html', {'entries': entries})


def archive_day(request, year, month, day):
    entries = BlogPost.objects.filter(draft=False,
                                      published_date__year=year,
                                      published_date__month=month,
                                      published_date__day=day)
    return render_to_response('BafosApp/index.html', {'entries': entries})


def details(request):
    entry = get_object_or_404(BlogPost)

    return render_to_response('BafosApp/details.html', {'entry': entry})


def youtube_parser(request):
    if 'text' in request.GET and request.GET['text']:
        channel_name = []
        video_count = []
        subscriber_count = []
        view_count = []
        channel_trailer = []
        channel_data = []
        creation_data = []
        request_text = request.GET.get('text', '')
        print(request_text)
        proxies = {
            "http": "http://45.250.226.56",
        }
        url = requests.get(str(request_text) + str('/page:') + str(1),proxies=proxies).text
        print(re.findall(r"http:\/\/\w{3}\.youtube\.com\/channel\/[a-zA-Z0-9-_-]+", url))
        page = 0
        while ('<strong>' in str(url)) == False:
            try:
                url = requests.get(str(request_text) + str('/page:') + str(page + 1),proxies=proxies).text
                soups = re.findall(r"http:\/\/\w{3}\.youtube\.com\/channel\/[a-zA-Z0-9-_-]+", url)
                for soup in soups:
                    if str(soup) not in channel_data:
                        print(soup)
                        channel_data.append(str(soup))

            except TimeoutError:
                print('error')
            page += 1
        for link in channel_data:
            time.sleep(3)
            response = requests.get(str('https://www.googleapis.com/youtube/v3/channels?'
                                        'part=snippet%2CcontentDetails%2Cstatistics%2CbrandingSettings&id=')
                                    + str(link).replace('http://www.youtube.com/channel/', '')
                                    + str('&key=AIzaSyACxrnyfBEZgUBNCwzCp7urOlORSzlZsHU'))
            data = response.json()
            try:
                channel_name.append(str(data['items'][0]['snippet']['title']))
            except:
                channel_name.append('Проблема с сетью или с каналом')
            try:
                video_count.append(str(data['items'][0]['statistics']['videoCount']))
            except:
                video_count.append('Проблема с сетью или с каналом')
            try:
                view_count.append(str(data['items'][0]['statistics']['viewCount']))
            except:
                view_count.append('Проблема с сетью или с каналом')
            try:
                subscriber_count.append(str(data['items'][0]['statistics']['subscriberCount']))
            except:
                subscriber_count.append('Проблема с сетью или с каналом')
            try:
                creation_data.append(str(data['items'][0]['snippet']['publishedAt'])[
                                 :str(data['items'][0]['snippet']['publishedAt']).find('T')])
            except:
                creation_data.append('Проблема с сетью или с каналом')
            try:
                channel_trailer.append(str('https://www.youtube.com/watch?v=') +
                                       str(data['items'][0]['brandingSettings']['channel']['unsubscribedTrailer']))
            except KeyError:
                channel_trailer.append('Трейлер канала отсутствует')
        return save_parser_data(channel_name, channel_data, video_count, view_count, subscriber_count, channel_trailer,
                                creation_data)

    return render_to_response('BafosApp/youtube_parser.html')


def save_parser_data(channel_name, channel_data, video_count, view_count, subscriber_count, channel_trailer, creation_data):
    table = pd.DataFrame({'Название канала': channel_name, 'Ссылка': channel_data, 'Количество видео': video_count,
                          'Общее количество просмотров': view_count, 'Количество подписчиков': subscriber_count,
                          'Трейлер канала': channel_trailer, 'Дата создания канала': creation_data},
                         columns=['Название канала', 'Ссылка', 'Дата создания канала', 'Количество видео',
                                  'Общее количество просмотров', 'Количество подписчиков', 'Трейлер канала'])

    table.to_csv("/home/" + str("channel_parser") + '.csv',
                 sep=';', index=False, encoding='utf-8-sig')
    excel_file_name = str("/home/") + str("channel_parser.csv")
    print("This is " + os.path.basename(excel_file_name))
    fp = open(excel_file_name, "rb");
    response = HttpResponse(fp.read());
    fp.close();
    file_type = mimetypes.guess_type(excel_file_name);
    if file_type is None:
        file_type = 'application/octet-stream';
    response['Content-Type'] = file_type
    response['Content-Length'] = str(os.stat(excel_file_name).st_size);
    response['Content-Disposition'] = "attachment; filename=%s" % os.path.basename(excel_file_name)
    return response


def get_adress(address):
    yandex_address = []
    street = []
    latitude = []
    longitude = []
    coordinate = []
    print(address)
    addr = []
    # for one_address in address:
    #   addr.append(str(one_address.split(",")[1])+","+str(one_address.split(",")[0]))
    #  print(addr)
    # address = addr
    for i in range(len(address)):
        response = requests.get(
            "https://geocode-maps.yandex.ru/1.x/?apikey=55080292-108d-4b98-b077-fb1c2af3affd&format=json&geocode=" + str(
                address[i])).json()
        print('tak')
        print(response)
        print(response['response']['GeoObjectCollection']['featureMember'][0]['GeoObject']['metaDataProperty'][
                  'GeocoderMetaData']['AddressDetails']['Country']['AddressLine'])
        try:
            # if 'Москва' in str(response['response']['GeoObjectCollection']['featureMember'][0]['GeoObject']['metaDataProperty']['GeocoderMetaData']['AddressDetails']['Country']['AddressLine']):
            print(str(
                response['response']['GeoObjectCollection']['featureMember'][0]['GeoObject']['metaDataProperty'][
                    'GeocoderMetaData']['AddressDetails']['Country']['AdministrativeArea']['Locality']))

            if 'Premise' in str(
                    response['response']['GeoObjectCollection']['featureMember'][0]['GeoObject']['metaDataProperty'][
                        'GeocoderMetaData']['AddressDetails']['Country']['AdministrativeArea']['Locality'][
                        'Thoroughfare']):
                print(str(
                    response['response']['GeoObjectCollection']['featureMember'][0]['GeoObject']['metaDataProperty'][
                        'GeocoderMetaData']['AddressDetails']['Country']['AdministrativeArea']['Locality'][
                        'Thoroughfare']['ThoroughfareName']) + ', ' + str(
                    response['response']['GeoObjectCollection']['featureMember'][0]['GeoObject']['metaDataProperty'][
                        'GeocoderMetaData']['AddressDetails']['Country']['AdministrativeArea']['Locality'][
                        'Thoroughfare']['Premise']['PremiseNumber']))
            else:
                print(str(response['response']['GeoObjectCollection']['featureMember'][0]['GeoObject'][
                              'metaDataProperty']['GeocoderMetaData']['AddressDetails']['Country'][
                              'AdministrativeArea']['Locality']['Thoroughfare']))

        except:
            if 'SubAdministrativeArea' in str(
                    response['response']['GeoObjectCollection']['featureMember'][0]['GeoObject'][
                        'metaDataProperty'][
                        'GeocoderMetaData']['AddressDetails']['Country']['AdministrativeArea']):
                if 'Thoroughfare' in str(response['response']['GeoObjectCollection']['featureMember'][0]['GeoObject'][
                                             'metaDataProperty'][
                                             'GeocoderMetaData']['AddressDetails']['Country']['AdministrativeArea'][
                                             'SubAdministrativeArea']['Locality']):
                    if 'Premise' in str(response['response']['GeoObjectCollection']['featureMember'][0]['GeoObject'][
                                            'metaDataProperty'][
                                            'GeocoderMetaData']['AddressDetails']['Country']['AdministrativeArea'][
                                            'SubAdministrativeArea']['Locality']['Thoroughfare']):
                        print(str(response['response']['GeoObjectCollection']['featureMember'][0]['GeoObject'][
                                      'metaDataProperty'][
                                      'GeocoderMetaData']['AddressDetails']['Country']['AdministrativeArea'][
                                      'SubAdministrativeArea']['Locality'][
                                      'Thoroughfare']['ThoroughfareName']) + ', ' + str(
                            response['response']['GeoObjectCollection']['featureMember'][0]['GeoObject'][
                                'metaDataProperty'][
                                'GeocoderMetaData']['AddressDetails']['Country']['AdministrativeArea'][
                                'SubAdministrativeArea']['Locality'][
                                'Thoroughfare']['Premise']['PremiseNumber']))
                    else:
                        print(str(response['response']['GeoObjectCollection']['featureMember'][0]['GeoObject'][
                                      'metaDataProperty'][
                                      'GeocoderMetaData']['AddressDetails']['Country']['AdministrativeArea'][
                                      'SubAdministrativeArea']['Locality'][
                                      'Thoroughfare']['ThoroughfareName']) + ', ' + str(
                            response['response']['GeoObjectCollection']['featureMember'][0]['GeoObject'][
                                'metaDataProperty'][
                                'GeocoderMetaData']['AddressDetails']['Country']['AdministrativeArea'][
                                'SubAdministrativeArea']['Locality'][
                                'Thoroughfare']))
                else:
                    print('Net Thoroughfare')
            else:
                print(str(response['response']['GeoObjectCollection']['featureMember'][0]['GeoObject'][
                              'metaDataProperty']['GeocoderMetaData']['AddressDetails']['Country'][
                              'AdministrativeArea']['Locality']))
                print(str(response['response']['GeoObjectCollection']['featureMember'][0]['GeoObject'][
                              'metaDataProperty']['GeocoderMetaData']['AddressDetails']['Country'][
                              'AdministrativeArea']['Locality']['DependentLocality']['DependentLocality'][
                              'DependentLocality']['DependentLocalityName']) + ', ' + str(
                    response['response']['GeoObjectCollection']['featureMember'][0]['GeoObject'][
                        'metaDataProperty']['GeocoderMetaData']['AddressDetails']['Country'][
                        'AdministrativeArea']['Locality']['DependentLocality']['DependentLocality'][
                        'DependentLocality']['Premise']['PremiseNumber']))

        print(str(response['response']['GeoObjectCollection']['featureMember'][0]['GeoObject']['Point']['pos']).split(
            " "))
        yandex_address.append(str(
            response['response']['GeoObjectCollection']['featureMember'][0]['GeoObject']['metaDataProperty'][
                'GeocoderMetaData']['AddressDetails']['Country']['AddressLine']))

        try:
            # if 'Москва' in str(response['response']['GeoObjectCollection']['featureMember'][0]['GeoObject']['metaDataProperty']['GeocoderMetaData']['AddressDetails']['Country']['AddressLine']):
            if 'Premise' in str(
                    response['response']['GeoObjectCollection']['featureMember'][0]['GeoObject']['metaDataProperty'][
                        'GeocoderMetaData']['AddressDetails']['Country']['AdministrativeArea']['Locality'][
                        'Thoroughfare']):
                print(str(
                    response['response']['GeoObjectCollection']['featureMember'][0]['GeoObject']['metaDataProperty'][
                        'GeocoderMetaData']['AddressDetails']['Country']['AdministrativeArea']['Locality'][
                        'Thoroughfare']))
                street.append(str(response['response']['GeoObjectCollection']['featureMember'][0]['GeoObject'][
                                      'metaDataProperty']['GeocoderMetaData']['AddressDetails']['Country'][
                                      'AdministrativeArea']['Locality']['Thoroughfare'][
                                      'ThoroughfareName']) + ', ' + str(
                    response['response']['GeoObjectCollection']['featureMember'][0]['GeoObject'][
                        'metaDataProperty']['GeocoderMetaData']['AddressDetails']['Country']['AdministrativeArea'][
                        'Locality']['Thoroughfare']['Premise']['PremiseNumber']))
            else:
                print(str(response['response']['GeoObjectCollection']['featureMember'][0]['GeoObject'][
                              'metaDataProperty']['GeocoderMetaData']['AddressDetails']['Country'][
                              'AdministrativeArea']['Locality']['Thoroughfare']))
                street.append(str(response['response']['GeoObjectCollection']['featureMember'][0]['GeoObject'][
                                      'metaDataProperty']['GeocoderMetaData']['AddressDetails']['Country'][
                                      'AdministrativeArea']['Locality']['Thoroughfare']['ThoroughfareName']))
        except:
            if 'SubAdministrativeArea' in str(
                    response['response']['GeoObjectCollection']['featureMember'][0]['GeoObject'][
                        'metaDataProperty'][
                        'GeocoderMetaData']['AddressDetails']['Country']['AdministrativeArea']):

                if 'Thoroughfare' in str(response['response']['GeoObjectCollection']['featureMember'][0]['GeoObject'][
                                             'metaDataProperty'][
                                             'GeocoderMetaData']['AddressDetails']['Country']['AdministrativeArea'][
                                             'SubAdministrativeArea']['Locality']):
                    if 'Premise' in str(
                            response['response']['GeoObjectCollection']['featureMember'][0]['GeoObject'][
                                'metaDataProperty'][
                                'GeocoderMetaData']['AddressDetails']['Country']['AdministrativeArea'][
                                'SubAdministrativeArea']['Locality']['Thoroughfare']):
                        street.append(str(response['response']['GeoObjectCollection']['featureMember'][0]['GeoObject'][
                                              'metaDataProperty'][
                                              'GeocoderMetaData']['AddressDetails']['Country']['AdministrativeArea'][
                                              'SubAdministrativeArea']['Locality'][
                                              'Thoroughfare']['ThoroughfareName']) + ', ' + str(
                            response['response']['GeoObjectCollection']['featureMember'][0]['GeoObject'][
                                'metaDataProperty'][
                                'GeocoderMetaData']['AddressDetails']['Country']['AdministrativeArea'][
                                'SubAdministrativeArea']['Locality'][
                                'Thoroughfare']['Premise']['PremiseNumber']))
                    else:
                        street.append(str(response['response']['GeoObjectCollection']['featureMember'][0]['GeoObject'][
                                              'metaDataProperty'][
                                              'GeocoderMetaData']['AddressDetails']['Country']['AdministrativeArea'][
                                              'SubAdministrativeArea']['Locality'][
                                              'Thoroughfare']['ThoroughfareName']) + ', ' + str(''))
                else:
                    street.append('')
            else:
                print(street.append(str(response['response']['GeoObjectCollection']['featureMember'][0]['GeoObject'][
                                            'metaDataProperty']['GeocoderMetaData']['AddressDetails']['Country'][
                                            'AdministrativeArea']['Locality']['DependentLocality']['DependentLocality'][
                                            'DependentLocality']['DependentLocalityName']) + ', ' + str(
                    response['response']['GeoObjectCollection']['featureMember'][0]['GeoObject'][
                        'metaDataProperty']['GeocoderMetaData']['AddressDetails']['Country'][
                        'AdministrativeArea']['Locality']['DependentLocality']['DependentLocality'][
                        'DependentLocality']['Premise']['PremiseNumber'])))

        coordinate.append(
            str(response['response']['GeoObjectCollection']['featureMember'][0]['GeoObject']['Point']['pos']).split(
                " "))
        print(coordinate)
        print(yandex_address)
        print(street)
        print(coordinate)
    print(coordinate)
    print(yandex_address)
    print(street)
    print(coordinate)
    for i in range(len(coordinate)):
        latitude.append(coordinate[i][1])
        longitude.append(coordinate[i][0])
    return yandex_address, street, latitude, longitude


def get_columns_data(path):
    try:
        wb = openpyxl.load_workbook(str(path))
    except FileNotFoundError:
        message = 'Укажите путь к таблице!'
        wb = openpyxl.load_workbook(str(path))
    sheet = wb.worksheets[0]
    address = []
    max_row = sheet.max_row
    for i in range(1, sheet.max_row):
        if sheet.cell(row=i, column=1).value is None:
            max_row = i - 1
            print(max_row)
            break
        else:
            max_row = sheet.max_row
            print(max_row)
    for i in range(2, max_row + 1):
        address.append(sheet.cell(row=i, column=1).value)
    return address


def get_vk_file():
    wb = openpyxl.load_workbook(str("/home/") + str("geocode.xlsx"))
    sheet = wb.worksheets[0]
    max_row = sheet.max_row
    for i in range(1, sheet.max_row):
        if ((sheet.cell(row=i, column=3).value) == None):
            max_row = i - 1
            break
        else:
            max_row = sheet.max_row

    column1 = []
    column2 = []
    column3 = []
    column4 = []

    for i in range(2, max_row + 1):
        column1.append(sheet.cell(row=i, column=3).value)
        column2.append(sheet.cell(row=i, column=4).value)
        column3.append(sheet.cell(row=i, column=5).value)
        column4.append(sheet.cell(row=i, column=6).value)

    keys = ('Улица', 'Радиус', 'Широта', 'Долгота')
    radius = (500, 1000, 1500, 2000, 2500, 3000, 3500, 4000, 4500, 5000, 6000,
              7000, 8000, 9000, 10000, 11000, 12000, 13000, 14000, 15000,
              20000, 25000, 30000, 35000, 40000, 45000, 50000, 55000,
              60000, 65000, 70000, 75000, 80000, 85000, 90000, 95000, 100000)

    for i in range(0, max_row - 1):
        column1[i] = column1[i].replace("улица", "ул.")
        column1[i] = column1[i].replace("площадь", "пл.")
        column1[i] = column1[i].replace("проспект", "пр-т")
        column1[i] = column1[i].replace("переулок", "пер.")
        column1[i] = column1[i].replace("проезд", "пр-д")

    ceil_column2 = []
    for var in column2:
        if int(var) in radius:
            ceil_column2.append(var)
        elif int(var) < 5000:
            if int(var) % 1000 < 500:
                ceil_column2.append((int(var) // 1000) * 1000 + 500)
            else:
                ceil_column2.append((int(var) // 1000) * 1000 + 1000)
        elif int(var) < 15000:
            ceil_column2.append(math.ceil(int(var) / 1000) * 1000)
        else:
            if int(var) % 10000 < 5000:
                ceil_column2.append(math.ceil(int(var) // 10000) * 10000 + 5000)
            else:
                ceil_column2.append(math.ceil(int(var) / 10000) * 10000)

    zipped = zip(column1, ceil_column2, column3, column4)

    dicts = [dict(zip(keys, values)) for values in zipped]

    final_list = []

    for i in range(0, max_row - 1):
        a = str(dicts[i]['Широта']) + ',' + str(dicts[i]['Долгота']) + ',' \
            + str(dicts[i]['Радиус']) + ',' + str(dicts[i]['Улица'])
        final_list.append(a)

    f = open(str("/home/") + 'VK' + '.txt', 'w',
             encoding="utf-8")

    for item in final_list:
        f.write("%s\n" % item)
    f.close()

    filename = str("/home/") + 'VK'
    return filename


def get_mytarget_file():
    wb = openpyxl.load_workbook(str("/home/") + str("geocode.xlsx"))
    sheet = wb.worksheets[0]
    max_row = sheet.max_row
    for i in range(1, sheet.max_row):
        if ((sheet.cell(row=i, column=3).value) == None):
            max_row = i - 1
            break
        else:
            max_row = sheet.max_row

    column1 = []
    column2 = []
    column3 = []
    column4 = []

    for i in range(2, max_row + 1):
        column1.append(sheet.cell(row=i, column=3).value)
        column2.append(sheet.cell(row=i, column=4).value)
        column3.append(sheet.cell(row=i, column=5).value)
        column4.append(sheet.cell(row=i, column=6).value)

    keys = ['Улица', 'Радиус', 'Широта', 'Долгота']

    for i in range(0, max_row - 1):
        column1[i] = column1[i].replace("улица", "ул.")
        column1[i] = column1[i].replace("площадь", "пл.")
        column1[i] = column1[i].replace("проспект", "пр-т")
        column1[i] = column1[i].replace("переулок", "пер.")
        column1[i] = column1[i].replace("проезд", "пр-д")

    for i in range(0, max_row - 1):
        column1[i] = column1[i]

    zipped = zip(column1, column2, column3, column4)

    dicts = [dict(zip(keys, values)) for values in zipped]

    final_list = []

    for i in range(0, max_row - 1):
        a = str(dicts[i]['Улица']) + ':' + str(dicts[i]['Радиус']) + 'm' + ':' + str(dicts[i]['Широта']) + ',' + str(
            dicts[i]['Долгота'])
        final_list.append(a)

    f = open(str("/home/") + 'MT' + '.txt', 'w',
             encoding="utf-8")

    for item in final_list:
        f.write("%s\n" % item)
    f.close()

    filename = str("/home/") + 'MT'
    return filename


def refill_table(path, yandex_adress, street, latitude, longitude, current_radius, current_service):
    wb = openpyxl.load_workbook(str(path))
    sheet = wb.worksheets[0]
    for i in range(0, len(yandex_adress)):
        print(yandex_adress[i])
        sheet.cell(row=i + 2, column=2).value = str(yandex_adress[i])
        sheet.cell(row=i + 2, column=3).value = str(street[i])
        sheet.cell(row=i + 2, column=4).value = str(current_radius)
        sheet.cell(row=i + 2, column=5).value = str(latitude[i])
        sheet.cell(row=i + 2, column=6).value = str(longitude[i])
    wb.save(str("/home/") + str("geocode.xlsx"))
    print(current_service)

    if current_service:
        for curr in current_service:
            print(curr)
            if curr == 'VK':
                filepath = get_vk_file()
                txt_vk_file_name = str(filepath) + str(".txt")
                fp = open(txt_vk_file_name, "rb")
                response = HttpResponse(fp.read())
                fp.close()
                file_type = mimetypes.guess_type(txt_vk_file_name)
                if file_type is None:
                    file_type = 'application/octet-stream'
                response['Content-Type'] = file_type
                response['Content-Length'] = str(os.stat(txt_vk_file_name).st_size)
                response['Content-Disposition'] = "attachment; filename=%s" % os.path.basename(txt_vk_file_name)
            elif curr == 'MyTarget':
                filepath = get_mytarget_file()
                excel_MT_file_name = str(filepath) + str(".txt")
                fp = open(excel_MT_file_name, "rb")
                response = HttpResponse(fp.read())
                fp.close()
                file_type = mimetypes.guess_type(excel_MT_file_name)
                if file_type is None:
                    file_type = 'application/octet-stream'
                response['Content-Type'] = file_type
                response['Content-Length'] = str(os.stat(excel_MT_file_name).st_size)
                response['Content-Disposition'] = "attachment; filename=%s" % os.path.basename(excel_MT_file_name)
    else:
        excel_file_name = str("/home/") + str("geocode.xlsx")
        fp = open(excel_file_name, "rb")
        response = HttpResponse(fp.read())
        fp.close()
        file_type = mimetypes.guess_type(excel_file_name)
        if file_type is None:
            file_type = 'application/octet-stream'
        response['Content-Type'] = file_type
        response['Content-Length'] = str(os.stat(excel_file_name).st_size)
        response['Content-Disposition'] = "attachment; filename=%s" % os.path.basename(excel_file_name)

    return response


def download_template(request):
    excel_file_name = str("media/") + str(request)
    fp = open(excel_file_name, "rb")
    response = HttpResponse(fp.read())
    fp.close()
    file_type = mimetypes.guess_type(excel_file_name)
    if file_type is None:
        file_type = 'application/octet-stream'
    response['Content-Type'] = file_type
    response['Content-Length'] = str(os.stat(excel_file_name).st_size)
    response['Content-Disposition'] = "attachment; filename=%s" % os.path.basename(excel_file_name)
    return response


def upload_file(request):
    print(request.method)
    print(request.POST)
    service = request.POST.getlist('checks')
    radius = request.POST.get('radius')
    print(service)
    print(radius)
    if request.method == 'POST':
        form = ArticleFileForm(request.POST, request.FILES)
        if len(service) != 0 and service[0] == 'template':
            return download_template('template.xlsx')
        else:
            if form.is_valid():
                try:
                    form.save()
                except:
                    print('test_error')
                return get_for_url_name(request.FILES['file'].name, radius, service)
            else:
                print(form.errors)
    else:
        form = ArticleFileForm()
    return render(request, 'BafosApp/uploaded.html', {'form': form})


def get_for_url_name(request, current_radius, current_service):
    path = str('media/') + str(request)
    #path = 'C:\\Users\\AMasanov\\media\\data.xlsx'
    address = get_columns_data(path)
    yandex_adress, street, latitude, longitude = get_adress(address)

    return refill_table(path, yandex_adress, street, latitude, longitude, current_radius, current_service)


def youtube(request):
    if 'text' in request.GET and request.GET['text']:
        text = request.GET.get('text', '')
        data = request.GET.get('data', '')

        request_text = text
        name_text = request_text
        video_name = []
        url_video = []
        description = []
        view_count = []
        vids = []

        id_channels_list = []
        description_list = []
        tittle = []
        view_count_channel = []
        comment_count = []
        subscriber_count = []
        video_count = []
        date_of_creation_of_channel = []
        country = []
        privacy_status = []
        keywords = []
        moderate_comments = []
        date_of_upload_latest_video = []
        if data == "video":
            response = requests.get('https://www.googleapis.com/youtube/v3/search?type=video'
                                '&key=AIzaSyACxrnyfBEZgUBNCwzCp7urOlORSzlZsHU&maxResults=50&'
                                'part=snippet&q=' + str(request_text))
            search_results = response.json()

            try:
                for i in range(len(search_results['items'])):
                    video_name.append(search_results['items'][i]['snippet']['title'])
                    url_video.append('https://www.youtube.com/watch?v=' + search_results['items'][i]['id']['videoId'])
                    description.append(search_results['items'][i]['snippet']['description'])
                    vids.append(search_results['items'][i]['id']['videoId'])
                if 'nextPageToken' in search_results:
                    try:
                        for i in range((int(search_results['pageInfo']['totalResults']) // 50) + 1):
                            #print('s is ' + str(search_results['pageInfo']['totalResults']))
                            if 'nextPageToken' in search_results:
                                response = requests.get('https://www.googleapis.com/youtube/v3/search?type=video'
                                                    '&key=AIzaSyACxrnyfBEZgUBNCwzCp7urOlORSzlZsHU&maxResults=50&'
                                                    '&part=snippet&pageToken=' + str(
                                        search_results['nextPageToken'])
                                                    + '&q=' + str(request_text))
                                search_results = response.json()
                                for j in range(len(search_results['items'])):
                                    video_name.append(search_results['items'][j]['snippet']['title'])
                                    url_video.append(
                                    'https://www.youtube.com/watch?v=' + search_results['items'][j]['id'][
                                        'videoId'])
                                    description.append(search_results['items'][j]['snippet']['description'])
                                    vids.append(search_results['items'][j]['id']['videoId'])
                    except:
                        print(i)
                        print('Something goes wrong. oops...')
            except:
                print('Видео не существует, или проблемы с сетью')

            if text != "":
                for vid in vids:
                    try:
                        video_response = requests.get('https://www.googleapis.com/youtube/v3/videos?part=statistics'
                                          '&key=AIzaSyACxrnyfBEZgUBNCwzCp7urOlORSzlZsHU&id=' + str(vid))
                        video_search_results = video_response.json()
                        print(video_search_results)
                        print(video_search_results['items'][0]['statistics']['viewCount'])
                        view_count.append(video_search_results['items'][0]['statistics']['viewCount'])
                    except:
                        view_count.append('N/A')

            table = pd.DataFrame({'Название видеоролика': video_name, 'Описание': description,
                                  'Количество просмотров': view_count, 'Url видеоролика': url_video})

            if text!="":
                table.to_csv("/home/" + str(datetime.datetime.now().strftime('%H-%M-%S_%d-%m-%Y'))+str("_video") + '.csv', sep=';', index=False,encoding='utf-8-sig')
                excel_file_name= str("/home/")+str(datetime.datetime.now().strftime('%H-%M-%S_%d-%m-%Y'))+str("_video.csv")
                print("This is " + os.path.basename(excel_file_name))
                fp = open(excel_file_name, "rb");
                response = HttpResponse(fp.read());
                fp.close();
                file_type = mimetypes.guess_type(excel_file_name);
                if file_type is None:
                    file_type = 'application/octet-stream';
                response['Content-Type'] = file_type
                response['Content-Length'] = str(os.stat(excel_file_name).st_size);
                response['Content-Disposition'] = "attachment; filename=%s" % os.path.basename(excel_file_name)
                return response;

            else:
                table.to_csv("/home/"+str(datetime.datetime.now().strftime('%H-%M-%S_%d-%m-%Y'))+str("_video") + '.csv', sep=';', index=False,
                             encoding='utf-8-sig')
                excel_file_name= str("/home/")+str(datetime.datetime.now().strftime('%H-%M-%S_%d-%m-%Y'))+str("_video.csv")
                print("This is " + os.path.basename(excel_file_name))
                fp = open(excel_file_name, "rb");
                response = HttpResponse(fp.read());
                fp.close();
                file_type = mimetypes.guess_type(excel_file_name);
                if file_type is None:
                    file_type = 'application/octet-stream';
                response['Content-Type'] = file_type
                response['Content-Length'] = str(os.stat(excel_file_name).st_size);
                response['Content-Disposition'] = "attachment; filename=%s" % os.path.basename(excel_file_name)
                return response;

        elif data == "channel":
            print('')
            print(request_text)
            response = requests.get('https://www.googleapis.com/youtube/v3/search?type=channel'
                                    '&key=AIzaSyACxrnyfBEZgUBNCwzCp7urOlORSzlZsHU&maxResults=50&'
                                    'part=snippet&q=' + str(request_text))
            search_results = response.json()
            print(search_results)
            try:
                for i in range(len(search_results['items'])):
                    print(search_results['items'][i]['snippet']['channelId'])
                    print(search_results['items'][i]['snippet']['description'])
                    id_channels_list.append(search_results['items'][i]['snippet']['channelId'])
                    description_list.append(search_results['items'][i]['snippet']['description'])
                    print('ne token')
                    if 'nextPageToken' in search_results:
                        try:
                            for i in range((int(search_results['pageInfo']['totalResults']) // 50) + 1):
                                print('s is ' + str(search_results['pageInfo']['totalResults']))
                                if 'nextPageToken' in search_results:
                                    print(i)
                                    print('token?')
                                    response = requests.get('https://www.googleapis.com/youtube/v3/search?type=channel'
                                                            '&key=AIzaSyACxrnyfBEZgUBNCwzCp7urOlORSzlZsHU&maxResults=50&'
                                                            '&part=snippet&pageToken=' + str(search_results['nextPageToken'])
                                                            + '&q=' + str(request_text))
                                    search_results = response.json()
                                    print(search_results)
                                    for j in range(len(search_results['items'])):
                                        print('token')
                                        print(search_results['items'][j]['snippet']['channelId'])
                                        print(search_results['items'][j]['snippet']['description'])
                                        id_channels_list.append(search_results['items'][j]['snippet']['channelId'])
                                        description_list.append(search_results['items'][j]['snippet']['description'])
                        except:
                            print(i)
                            print('Something goes wrong. oops...')
            except:
                print('Канала не существует, или проблемы с сетью')

            for name in id_channels_list:
                content = requests.get('https://www.googleapis.com/youtube/v3/channels?id=' + str(name) + '&key=AIzaSyACxrnyfBEZgUBNCwzCp7urOlORSzlZsHU&part=snippet,statistics,status,brandingSettings')
                data = content.json()
                print(" ")
                print(data["items"][0]["snippet"]["title"])
                print(data["items"][0]["statistics"]["viewCount"])
                print(data["items"][0]["statistics"]["commentCount"])
                print(data["items"][0]["statistics"]["subscriberCount"])
                print(data["items"][0]["statistics"]["videoCount"])
                tittle.append(data["items"][0]["snippet"]["title"])
                view_count_channel.append(data["items"][0]["statistics"]["viewCount"])
                comment_count.append(data["items"][0]["statistics"]["commentCount"])
                subscriber_count.append(data["items"][0]["statistics"]["subscriberCount"])
                video_count.append(data["items"][0]["statistics"]["videoCount"])
                date_of_creation_of_channel.append(data["items"][0]["snippet"]["publishedAt"])
                try:
                    country.append(data["items"][0]["snippet"]["country"])
                except:
                    country.append("None")

                try:
                    privacy_status.append(data["items"][0]["status"]["privacyStatus"])
                except:
                    privacy_status.append("None")

                try:
                    keywords.append(data["items"][0]["brandingSettings"]["channel"]["keywords"])
                except:
                    keywords.append("None")

                try:
                    moderate_comments.append(data["items"][0]["brandingSettings"]["channel"]["moderateComments"])
                except:
                    moderate_comments.append("None")

            for name in id_channels_list:
                try:
                    response = requests.get('https://www.googleapis.com/youtube/v3/activities?channelId=' + str(name) + '&key=AIzaSyACxrnyfBEZgUBNCwzCp7urOlORSzlZsHU&part=snippet')
                    last_upload = response.json()
                    date_of_upload_latest_video.append(last_upload["items"][0]["snippet"]["publishedAt"])
                except:
                    date_of_upload_latest_video.append("None")

            channel_urls = []
            for channel_url in id_channels_list:
                channel_urls.append("https://www.youtube.com/channel/"+str(channel_url))

            table = pd.DataFrame({'Ссылка на канал':channel_urls,'Название канала': tittle, 'Описание канала': description_list,
            'Общее количество просмотров': view_count_channel,
            'Количество подписчиков': subscriber_count, 'Количество видеороликов': video_count,
            'Дата создания канала':date_of_creation_of_channel, 'Страна':country,
            'Доступность канала':privacy_status, 'Ключевые слова': keywords,'Модерация комментариев': moderate_comments,
            'Дата загрузки последного видеоролика': date_of_upload_latest_video})

            if text!="":
                table.to_csv("/home/"+str(datetime.datetime.now().strftime('%H-%M-%S_%d-%m-%Y'))+ str("_channel") + '.csv', sep=';', index=False,
                             encoding='utf-8-sig')
                excel_file_name= str("/home/")+str(datetime.datetime.now().strftime('%H-%M-%S_%d-%m-%Y'))+str("_channel.csv")
                fp = open(excel_file_name, "rb");
                response = HttpResponse(fp.read());
                fp.close();
                file_type = mimetypes.guess_type(excel_file_name);
                if file_type is None:
                    file_type = 'application/octet-stream';
                response['Content-Type'] = file_type
                response['Content-Length'] = str(os.stat(excel_file_name).st_size);
                response['Content-Disposition'] = "attachment; filename=%s" % os.path.basename(excel_file_name)
                return response;

            else:
                table.to_csv("/home/"+str(datetime.datetime.now().strftime('%H-%M-%S_%d-%m-%Y'))+str("_channel") + '.csv', sep=';', index=False,
                             encoding='utf-8-sig')
                excel_file_name= str("/home/")+str(datetime.datetime.now().strftime('%H-%M-%S_%d-%m-%Y'))+str("_channel.csv")
                fp = open(excel_file_name, "rb");
                response = HttpResponse(fp.read());
                fp.close();
                file_type = mimetypes.guess_type(excel_file_name);
                if file_type is None:
                    file_type = 'application/octet-stream';
                response['Content-Type'] = file_type
                response['Content-Length'] = str(os.stat(excel_file_name).st_size);
                response['Content-Disposition'] = "attachment; filename=%s" % os.path.basename(excel_file_name)
                return response;

        else:
            print('')

    text = 'tratata'
    data = 'kek'
    context = {
        'text': "CCHORT",
        'data': "DIRT",
    }

    if text and data:
        print('Text: {}\nData: {}'.format("CCHORT", "DIRT"))

    return render_to_response('BafosApp/youtube.html')


def ads_names(request):

    return render_to_response('BafosApp/ads_names.html')

def mail_naming(request):

    return render_to_response('BafosApp/mail_naming.html')


regions = {'Москва': 'Московская область', 'Санкт-Петербург': 'Ленинградская область',
           'Нижний Новгород': 'Нижегородская область', 'Ростов-на-Дону':'Ростовская область',
           'Екатеринбург': 'Свердловская область', 'Владимир': 'Владимирская область',
           'Великий Новгород': 'Новогородская область', 'Рязань': 'Рязанская область', 'Тюмень': 'Тюменская область',
           'Липецк': 'Липецкая область', 'Тула': 'Тульская область', 'Вологода': 'Вологодская область',
           'Орёл': 'Орловская область', 'Калуга': 'Калужская область', 'Смоленск': 'Смоленская область',
           'Архангельск': 'Архангельская область', 'Тверь': 'Тверская область', 'Брянск': 'Брянская область',
           'Белгород': 'Белгородская область', 'Ярославль': 'Ярославская область', 'Астрахань': 'Астраханская область',
           'Тамбов': 'Тамбовская область', 'Калининград': 'Калининградская область', 'Курск': 'Курская область',
           'Кострома': 'Костромская область', 'Курган': 'Курганская область', 'Мурманск': 'Мурманская область',
           'Иркутск': 'Иркутская область', 'Новосибирск': 'Новосибирская область', 'Магадан': 'Магаданская область',
           'Барнаул': 'Алтаский край', 'Воронеж': 'Воронежская область', 'Иваново': 'Ивановская область',
           'Майкоп': 'Республика Адыгея', 'Элиста': 'Республика Калмыкия', 'Краснодар': 'Краснодарский край',
           'Волгоград': 'Волгоградская область', 'Петрозаводск': 'Республика Карелия', 'Сыктывкар': 'Республика Коми',
           'Псков': 'Псковская область', 'Нарьян-Мар': 'Ненецкий автономный округ', 'Якутск': 'Республика Саха',
           'Петропавловск-Камчатский': 'Камчатский край', 'Владивосток': 'Приморский край',
           'Хабаровск': 'Хабаровский край', 'Благовещенск': 'Амурская область', 'Южно-Сахалинск': 'Сахалинская область',
           'Биробиджан': 'Еврейская автономная область', 'Анадырь': 'Чукотский автономный округ',
           'Горно-Алтайск': 'Республика Алтай', 'Улан-Удэ': 'Республика Бурятия', 'Кызыл': 'Республика Тыва',
           'Абакан': 'Республика Хакасия', 'Чита': 'Забайкальский край', 'Красноярск': 'Красноярский край',
           'Кемерово': 'Кемеровская область', 'Омск': 'Омская область', 'Томск': 'Томская область',
           'Челябинск': 'Челябинская область', 'Ханты-Мансийск': 'Ханты-Мансийский автономный округ - Югра',
           'Салехард': 'Ямало-Ненецкий автономный округ', 'Уфа': 'Республика Башкортостан',
           'Йошкар-Ола': 'Республика Марий Эл', 'Саранск': 'Республика Мордовия', 'Казань': 'Республика Татарстан',
           'Ижевск': 'Удмуртская Республика', 'Чебоксары': 'Чувашская Республика', 'Киров': 'Кировская область',
           'Оренбург': 'Оренбургская область', 'Пенза': 'Пензенская область', 'Ульяновск': 'Ульяновская область',
           'Самара': 'Самарская область', 'Саратов': 'Саратовская область', 'Пермь': 'Пермский край',
           'Махачкала': 'Республика Дагестан', 'Магас': 'Республика Ингушетия',
           'Нальчик': 'Кабардино-Балкарская Республика'}


def get_full_adress(organization, result):
    offset = 0
    address = []
    coordinates_list = []
    organization_list = []

    url = requests.get(
        'https://search-maps.yandex.ru/v1/?text=' + str(organization) + '&results=' + str(result) +
        '&skip=' + str(offset) + '&rspn=1&type=biz&lang=ru_RU&' + 'apikey=0ca10b46-53c4-49bd-9c85-74bca3d66127')
    response = url.json()
    value = response['properties']['ResponseMetaData']['SearchResponse']['found']
    print(value)
    for item in response['features']:
        if organization.count(',') == 2:
            organization = organization.split(',')[2]
        elif organization.count(',') == 1:
            organization = organization.split(',')[1]

        if organization.startswith(" "):
            organization = organization[1:]

        organization_list.append(organization)
        if item['properties']['CompanyMetaData']['address'] not in address:
            print(item['properties']['CompanyMetaData']['address'])
            address.append(item['properties']['CompanyMetaData']['address'])

        if item['geometry']['coordinates'] not in coordinates_list:
            coordinates_list.append(item['geometry']['coordinates'])

    print(len(address))
    return address, coordinates_list, organization_list


def get_address(request, organization, result):
    offset = 0
    address = []
    coordinates_list = []
    organization_list = []

    url = requests.get('https://search-maps.yandex.ru/v1/?text=' + str(organization) + '&results=' + str(result) + '&skip=' + str(offset) + '&rspn=1&type=biz&lang=ru_RU&' + 'apikey=0ca10b46-53c4-49bd-9c85-74bca3d66127')
    response = url.json()
    value = response['properties']['ResponseMetaData']['SearchResponse']['found']
    print(value)

    for item in response['features']:
        if organization.count(',') == 2:
            organization = organization.split(',')[2]
        elif organization.count(',') == 1:
            organization = organization.split(',')[1]

        if organization.startswith(" "):
            organization = organization[1:]

        organization_list.append(organization)
    meta_address = []
    for item in response['features']:
        if item['properties']['CompanyMetaData']['address'] not in meta_address:
            print(item['properties']['CompanyMetaData']['address'])
            meta_address.append(item['properties']['CompanyMetaData']['address'])

        if item['geometry']['coordinates'] not in coordinates_list:
            coordinates_list.append(item['geometry']['coordinates'])

    if request.is_ajax():
        print('работаем')
        return HttpResponse('pending')
    address = example.get_address(request,meta_address)

    return address, coordinates_list, organization_list


def create_only_address_list(dicts, address_list):
    final_result = []
    try:
        for i in range(0, len(address_list)):
            element = str(dicts[i]['Улица'])
            final_result.append(element)
    except:
        pass
    return final_result


def create_final_list(dicts, address_list):
    final_result = []
    try:
        for i in range(0, len(address_list)):
            element = str(dicts[i]['Улица']) + ':' + str(dicts[i]['Радиус']) + 'm' + ':' + str(dicts[i]['Широта']) + ',' + str(dicts[i]['Долгота'])
            final_result.append(element)
    except:
        pass
    return final_result


def save_file(final_result):
    filetype = [('Text File', '.txt')]
    rer = str("/home/")
    f = open(rer +'file_' + str(datetime.datetime.now().strftime('%H-%M-%S_%d-%m-%Y')) + '.txt', 'w', encoding="utf-8")
    # Save list into file

    # Path_to_file+str(datetime.today().strftime("%Y-%m-%d-%H.%M"))+'.txt'
    for item in final_result:
        f.write("%s\n" % item)
    f.close()
    excel_file_name= str("/home/") + str("file_")+str(datetime.datetime.now().strftime('%H-%M-%S_%d-%m-%Y'))+str(".txt")
    print("This is " + os.path.basename(excel_file_name))
    fp = open(excel_file_name, "rb");
    response = HttpResponse(fp.read());
    fp.close();
    file_type = mimetypes.guess_type(excel_file_name);
    if file_type is None:
        file_type = 'application/octet-stream';
    response['Content-Type'] = file_type
    response['Content-Length'] = str(os.stat(excel_file_name).st_size);
    response['Content-Disposition'] = "attachment; filename=file.txt";
    return response;


def list_to_dict(name_of_organization, address_list, radius_list, latitude, longitude):
    keys = ['Организация', 'Улица', 'Радиус', 'Широта', 'Долгота']
    zipped = zip(name_of_organization, address_list, radius_list, latitude, longitude)
    new_dicts = [dict(zip(keys, values)) for values in zipped]
    return new_dicts


def rename_objects(address_list,var):
    for i in range(0, len(address_list)):
        address_list[i] = address_list[i].replace("улица", "ул.")
        address_list[i] = address_list[i].replace("площадь", "пл.")
        address_list[i] = address_list[i].replace("проспект", "пр-т")
        address_list[i] = address_list[i].replace("просп.", "пр-т")
        address_list[i] = address_list[i].replace("переулок", "пер.")
        address_list[i] = address_list[i].replace("проезд", "пр-д")
        address_list[i] = address_list[i].replace("Центральный федеральный округ", "ЦФО")
        if not var == 1:
            address_list[i] = address_list[i][1:]
    return address_list


def sort_coordinate(name_of_organization, address_list, coordinate, radius):
    # Break coordinate into two lists, latitude and longitude
    latitude = []
    longitude = []

    # -------------------------------------------------------------------------------------------
    # Create radius list for test, after publication it will be get values from excel sheet
    radius_list = []
    for rad in range(0, len(address_list)):
        radius_list.append(radius)
    # -------------------------------------------------------------------------------------------
    for sort_coordinate in coordinate:
        latitude.append(sort_coordinate[1])
        longitude.append(sort_coordinate[0])
    return radius_list, latitude, longitude


@csrf_exempt
def organizations(request):
    if request.is_ajax():
        print('работаем')
        return HttpResponse('pending')

    if 'сompany' in request.POST and request.POST['сompany']:
        company = request.POST.get('сompany', '')
        radius = request.POST.get('radius', '')
        data = request.POST.get('data', '')
        enable_cities = request.POST.getlist('checks[]')
        enable_regions = request.POST.getlist('checks_dist[]')
        print(enable_cities)
        print(enable_regions)
        address_list = []
        coordinate = []
        name_of_organization = []
        if data == 'Выгрузить полные адреса':
            var = 1
            var_only_address = 0
        else:
            var = 0
            var_only_address = 1

        try:
            business = company
            radius_count = radius
            address_list = []
            coordinate = []
            name_of_organization = []
            for key_value in regions:
                #time.sleep(1)
                if key_value in enable_cities and regions[key_value] in enable_regions:
                    print(key_value)
                    if var == 1:
                        cities_address_list, cities_coordinate, cities_name_of_organization = get_full_adress(key_value + ',' + business, result=1000)
                        if cities_address_list not in address_list:
                            address_list = address_list + cities_address_list
                        if cities_coordinate not in coordinate:
                            coordinate = coordinate + cities_coordinate
                        if cities_name_of_organization not in name_of_organization:
                            name_of_organization = name_of_organization + cities_name_of_organization

                        print(regions[key_value])
                        regions_address_list, regions_coordinate, regions_name_of_organization = get_full_adress(regions[key_value] + ',' + business, result=1000)
                        if regions_address_list not in address_list:
                            address_list = address_list + regions_address_list
                        if regions_coordinate not in coordinate:
                            coordinate = coordinate + regions_coordinate
                        if regions_name_of_organization not in name_of_organization:
                            name_of_organization = name_of_organization + regions_name_of_organization

                    else:
                        cities_address_list, cities_coordinate, cities_name_of_organization = get_address(request, key_value +',' + business, result=1000)
                        if cities_address_list not in address_list:
                            address_list = address_list + cities_address_list
                        if cities_coordinate not in coordinate:
                            coordinate = coordinate + cities_coordinate
                        if cities_name_of_organization not in name_of_organization:
                            name_of_organization = name_of_organization + cities_name_of_organization

                        print(regions[key_value])
                        regions_address_list, regions_coordinate, regions_name_of_organization = get_address(request, regions[key_value] + ',' + business, result=1000)
                        if regions_address_list not in address_list:
                            address_list = address_list + regions_address_list
                        if regions_coordinate not in coordinate:
                            coordinate = coordinate + regions_coordinate
                        if regions_name_of_organization not in name_of_organization:
                            name_of_organization = name_of_organization + regions_name_of_organization

                elif key_value in enable_cities:
                    print(key_value)
                    if var == 1:
                        old_address_list, old_coordinate, old_name_of_organization = get_full_adress(key_value + ',' + business, result=1000)
                        if old_address_list not in address_list:
                            address_list = address_list + old_address_list
                        if old_coordinate not in coordinate:
                            coordinate = coordinate + old_coordinate
                        if old_name_of_organization not in name_of_organization:
                            name_of_organization = name_of_organization + old_name_of_organization

                    else:
                        old_address_list, old_coordinate, old_name_of_organization = get_address(request,key_value+',' + business, result=1000)
                        if old_address_list not in address_list:
                            address_list = address_list + old_address_list
                        if old_coordinate not in coordinate:
                            coordinate = coordinate + old_coordinate
                        if old_name_of_organization not in name_of_organization:
                            name_of_organization = name_of_organization + old_name_of_organization

                elif regions[key_value] in enable_regions:
                    print(regions[key_value])
                    if var == 1:
                        old_address_list, old_coordinate, old_name_of_organization = get_full_adress(regions[key_value] + ',' + business, result=1000)
                        if old_address_list not in address_list:
                            address_list = address_list + old_address_list
                        if old_coordinate not in coordinate:
                            coordinate = coordinate + old_coordinate
                        if old_name_of_organization not in name_of_organization:
                            name_of_organization = name_of_organization + old_name_of_organization
                    else:
                        old_address_list, old_coordinate, old_name_of_organization = get_address(request, regions[key_value] + ',' + business, result=1000)
                        if old_address_list not in address_list:
                            address_list = address_list + old_address_list
                        if old_coordinate not in coordinate:
                            coordinate = coordinate + old_coordinate
                        if old_name_of_organization not in name_of_organization:
                            name_of_organization = name_of_organization + old_name_of_organization

            address_list = rename_objects(address_list,var)
            radius_list, latitude, longitude = sort_coordinate(name_of_organization, address_list, coordinate, radius_count)
            dicts = list_to_dict(name_of_organization, address_list, radius_list, latitude, longitude)

            if var_only_address == 1:
                final_result = create_only_address_list(dicts,address_list)
            elif not var_only_address == 1:
                final_result = create_final_list(dicts, address_list)

            #path_to_file = os.path.realpath('file_' + str(datetime.today().strftime("%Y-%m-%d-%H.%M")))
            #save_file(final_result)
            if final_result != []:
                filetype = [('Text File', '.txt')]
                rer = str("/home/")
                f = open(rer +'file_' + str(datetime.datetime.now().strftime('%H-%M-%S_%d-%m-%Y')) + '.txt', 'w', encoding="utf-8")
                # Save list into file

                # Path_to_file+str(datetime.today().strftime("%Y-%m-%d-%H.%M"))+'.txt'
                for item in final_result:
                    f.write("%s\r\n" % item)
                f.close()
                excel_file_name= str("/home/") + str("file_")+str(datetime.datetime.now().strftime('%H-%M-%S_%d-%m-%Y'))+str(".txt")
                print("This is " + os.path.basename(excel_file_name))
                fp = open(excel_file_name, "rb");
                response = HttpResponse(fp.read());
                fp.close();
                file_type = mimetypes.guess_type(excel_file_name);
                if file_type is None:
                    file_type = 'application/octet-stream';
                response['Content-Type'] = file_type
                response['Content-Length'] = str(os.stat(excel_file_name).st_size);
                response['Content-Disposition'] = "attachment; filename=%s" % os.path.basename(excel_file_name)
                os.remove(excel_file_name)
                return response;

        except IndexError:
            message = 'Индекс вышел за границы массива'
        except UnboundLocalError:
            message = 'Выберите требуемый город/область'
        except ValueError:
            message = 'Обнаруженые пустые поля/Неверный тип переменной'

    return render_to_response('BafosApp/organizations.html')